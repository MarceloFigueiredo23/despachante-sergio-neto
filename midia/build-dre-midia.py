# -*- coding: utf-8 -*-
"""DRE projetado de mídia — Despachante Sérgio Neto."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "DRE-midia-despachante-sergio-neto.xlsx"

# --- Dados da planilha de custos do cliente ---
SERVICOS = [
    ("Transferência", 296.0, 4.70, 300.70, 510.0, 209.30),
    ("Licenciamento", 175.0, 4.70, 179.70, 290.0, 110.30),
    ("Licenciamento + Transferência", 471.0, 4.70, 475.70, 690.0, 214.30),
    ("Vistoria", 90.0, 0.0, 90.0, 130.0, 40.0),
    ("Placa Moto", 85.0, 0.0, 85.0, 150.0, 65.0),
    ("Placa Carro", 140.0, 0.0, 140.0, 245.0, 105.0),
    ("APTV", 4.70, 0.0, 4.70, 45.0, 40.30),
    ("Emissão de Porte", 4.70, 0.0, 4.70, 45.0, 40.30),
]

# Mix esperado de serviços vindos de tráfego pago (ajustável)
MIX = {
    "Transferência": 0.35,
    "Licenciamento": 0.25,
    "Licenciamento + Transferência": 0.20,
    "Placa Carro": 0.10,
    "Vistoria": 0.05,
    "APTV": 0.05,
}

lucro_by_name = {s[0]: s[5] for s in SERVICOS}
valor_by_name = {s[0]: s[4] for s in SERVICOS}
LUCRO_MEDIO = sum(lucro_by_name[k] * w for k, w in MIX.items())
TICKET_MEDIO = sum(valor_by_name[k] * w for k, w in MIX.items())

# Premissas de performance (local Cosmópolis — conservador)
# CPA = custo por serviço FECHADO (não só lead)
META = {
    "nome": "Meta Ads (Instagram/Facebook + WhatsApp)",
    "cpa_conservador": 75.0,
    "cpa_base": 55.0,
    "cpa_otimista": 40.0,
    "pct_orcamento_500": 0.70,
    "pct_orcamento_1000": 0.60,
}
GOOGLE = {
    "nome": "Google Ads (Search + Maps)",
    "cpa_conservador": 95.0,
    "cpa_base": 70.0,
    "cpa_otimista": 50.0,
    "pct_orcamento_500": 0.30,
    "pct_orcamento_1000": 0.40,
}


def style_header(cell, fill="1A3358"):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_money(cell):
    cell.number_format = 'R$ #,##0.00'
    cell.font = Font(name="Arial", size=10)


def style_pct(cell):
    cell.number_format = "0.0%"
    cell.font = Font(name="Arial", size=10)


def thin_border(ws, r1, c1, r2, c2):
    b = Border(
        left=Side(style="thin", color="D0D7E2"),
        right=Side(style="thin", color="D0D7E2"),
        top=Side(style="thin", color="D0D7E2"),
        bottom=Side(style="thin", color="D0D7E2"),
    )
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = b


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def vendas_por_canal(orcamento, canal, cenario="base"):
    key = f"cpa_{cenario}"
    pct_key = "pct_orcamento_500" if orcamento <= 500 else "pct_orcamento_1000"
    spend = orcamento * canal[pct_key]
    cpa = canal[key]
    vendas = spend / cpa
    return spend, cpa, vendas


def montar_cenario(orcamento, cenario="base"):
    sm, cpa_m, vm = vendas_por_canal(orcamento, META, cenario)
    sg, cpa_g, vg = vendas_por_canal(orcamento, GOOGLE, cenario)
    vendas = vm + vg
    receita = vendas * TICKET_MEDIO
    lucro_bruto = vendas * LUCRO_MEDIO
    lucro_apos_midia = lucro_bruto - orcamento
    roi = (lucro_apos_midia / orcamento) if orcamento else 0
    be_vendas = orcamento / LUCRO_MEDIO
    return {
        "orcamento": orcamento,
        "spend_meta": sm,
        "spend_google": sg,
        "cpa_meta": cpa_m,
        "cpa_google": cpa_g,
        "vendas_meta": vm,
        "vendas_google": vg,
        "vendas": vendas,
        "receita": receita,
        "lucro_bruto": lucro_bruto,
        "lucro_apos_midia": lucro_apos_midia,
        "roi": roi,
        "be_vendas": be_vendas,
        "roas_lucro": lucro_bruto / orcamento if orcamento else 0,
    }


def sheet_capa(wb):
    ws = wb.active
    ws.title = "00_Resumo"
    ws["A1"] = "DRE PROJETADO DE MÍDIA — Despachante Sérgio Neto (Biju)"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="1A3358")
    ws.merge_cells("A1:G1")
    ws["A2"] = "Cosmópolis-SP · Campo do Kigol · Projeção para investimento em tráfego pago"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="5A6D86")

    rows = [
        ("", ""),
        ("Lucro bruto médio ponderado (mix tráfego)", LUCRO_MEDIO),
        ("Ticket médio ponderado (valor cliente)", TICKET_MEDIO),
        ("", ""),
        ("VEREDICTO RÁPIDO", ""),
        ("Melhor canal principal", "Meta Ads (WhatsApp) — volume local"),
        ("Melhor canal de intenção", "Google Search/Maps — quem já busca despachante"),
        ("Mix sugerido em R$ 500", "70% Meta · 30% Google"),
        ("Mix sugerido em R$ 1.000", "60% Meta · 40% Google"),
        ("Raio de targeting", "15–20 km (Cosmópolis + Artur Nogueira)"),
        ("Faixa etária", "25–54 anos"),
        ("LT break-even (cenário base)", "Mês 1 — se bater CPA base"),
        ("LT início de ROI consistente", "Mês 2–3 (após otimização)"),
    ]
    r = 4
    for label, val in rows:
        ws.cell(r, 1, label).font = Font(name="Arial", bold=bool(label and val == "" or "VEREDICTO" in str(label) or "Melhor" in str(label) or "Mix" in str(label) or "LT" in str(label) or "Raio" in str(label) or "Faixa" in str(label)), size=11, color="1A3358")
        cell = ws.cell(r, 2, val)
        cell.font = Font(name="Arial", size=11)
        if isinstance(val, float):
            style_money(cell)
        r += 1

    # mini table 500 vs 1000
    r += 1
    ws.cell(r, 1, "COMPARAÇÃO RÁPIDA — CENÁRIO BASE").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    headers = ["Métrica", "R$ 500/mês", "R$ 1.000/mês"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(r, i, h))
    c500 = montar_cenario(500, "base")
    c1000 = montar_cenario(1000, "base")
    metrics = [
        ("Investimento mídia", c500["orcamento"], c1000["orcamento"]),
        ("Vendas estimadas / mês", c500["vendas"], c1000["vendas"]),
        ("Vendas p/ break-even", c500["be_vendas"], c1000["be_vendas"]),
        ("Receita bruta estimada", c500["receita"], c1000["receita"]),
        ("Lucro bruto serviços", c500["lucro_bruto"], c1000["lucro_bruto"]),
        ("Lucro após mídia", c500["lucro_apos_midia"], c1000["lucro_apos_midia"]),
        ("ROI sobre mídia", c500["roi"], c1000["roi"]),
        ("ROAS (lucro/invest.)", c500["roas_lucro"], c1000["roas_lucro"]),
    ]
    for i, (name, a, b) in enumerate(metrics, 1):
        rr = r + i
        ws.cell(rr, 1, name).font = Font(name="Arial", size=10)
        ca, cb = ws.cell(rr, 2, a), ws.cell(rr, 3, b)
        if "ROI" in name or "ROAS" in name:
            style_pct(ca)
            style_pct(cb)
        elif "Vendas" in name and "estimadas" in name or "break-even" in name:
            ca.number_format = "0.0"
            cb.number_format = "0.0"
            ca.font = Font(name="Arial", size=10)
            cb.font = Font(name="Arial", size=10)
        else:
            style_money(ca)
            style_money(cb)
        if "Lucro após" in name:
            ca.font = Font(name="Arial", bold=True, color="1A7F37", size=10)
            cb.font = Font(name="Arial", bold=True, color="1A7F37", size=10)

    ws.cell(r + len(metrics) + 2, 1, "Obs.: projeções usam CPA médio local (cidade pequena). Validar nas 2 primeiras semanas e ajustar.").font = Font(name="Arial", italic=True, size=9, color="5A6D86")
    autosize(ws, [42, 22, 22, 18])


def sheet_premissas(wb):
    ws = wb.create_sheet("01_Premissas_Custos")
    ws["A1"] = "Tabela de custos e lucro (fonte: planilha do negócio)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3358")
    headers = ["Serviço", "Custo Processo", "Custo Operacional", "Custo Total", "Valor Cliente", "Lucro Bruto", "Mix tráfego", "Lucro ponderado"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h), "1A3358")

    r = 4
    for nome, cp, co, ct, vc, lb in SERVICOS:
        ws.cell(r, 1, nome).font = Font(name="Arial", size=10)
        for c, v in enumerate([cp, co, ct, vc, lb], 2):
            style_money(ws.cell(r, c, v))
        mix = MIX.get(nome, 0)
        style_pct(ws.cell(r, 7, mix))
        style_money(ws.cell(r, 8, lb * mix))
        r += 1

    r += 1
    ws.cell(r, 1, "LUCRO MÉDIO PONDERADO").font = Font(name="Arial", bold=True, size=11)
    style_money(ws.cell(r, 2, LUCRO_MEDIO))
    ws.cell(r, 2).font = Font(name="Arial", bold=True, color="1A7F37", size=12)
    r += 1
    ws.cell(r, 1, "TICKET MÉDIO PONDERADO").font = Font(name="Arial", bold=True, size=11)
    style_money(ws.cell(r, 2, TICKET_MEDIO))

    r += 3
    ws.cell(r, 1, "Como ler").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    ws.cell(r, 1, "O lucro bruto é o que sobra DEPOIS dos custos do processo e operacionais — é ele que paga a mídia e gera ROI.").font = Font(name="Arial", size=10)
    r += 1
    ws.cell(r, 1, "Break-even de mídia = Investimento ÷ Lucro bruto médio. Ex.: R$ 500 ÷ R$ 158 ≈ 3,2 serviços fechados.").font = Font(name="Arial", size=10)
    r += 1
    ws.cell(r, 1, "Ajuste a coluna Mix se a carteira de leads vier diferente (ex.: mais licenciamento no começo do ano).").font = Font(name="Arial", size=10)

    thin_border(ws, 3, 1, 3 + len(SERVICOS), 8)
    autosize(ws, [32, 16, 16, 14, 14, 14, 12, 16])


def sheet_canais(wb):
    ws = wb.create_sheet("02_Canais_Targeting")
    ws["A1"] = "Google vs Meta — o que usar, raio, idade e criativos"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3358")
    ws.merge_cells("A1:F1")

    headers = ["Tema", "Meta Ads", "Google Ads", "Recomendação"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h))

    data = [
        (
            "Papel principal",
            "Volume + WhatsApp (topo/meio de funil)",
            "Intenção alta (fundo de funil)",
            "Meta como motor; Google como captura",
        ),
        (
            "Formatos",
            "Tráfego p/ WhatsApp, Local Awareness, Reels/Feed",
            "Search + Performance Max + Perfil Business/Maps",
            "WhatsApp clique no Meta; palavras-chave no Google",
        ),
        (
            "CPA estimado (base)",
            f"R$ {META['cpa_base']:.0f} / serviço fechado",
            f"R$ {GOOGLE['cpa_base']:.0f} / serviço fechado",
            "Meta mais barato; Google converte melhor quem busca",
        ),
        (
            "Quando priorizar",
            "Orçamento baixo, marca nova, prova social",
            "Já tem demanda de busca / Maps forte",
            "Começar Meta; escalar Google no R$ 1.000",
        ),
        (
            "Raio geográfico",
            "15–20 km do Campo do Kigol",
            "Cosmópolis + Artur Nogueira (+ interesse)",
            "Não abrir >25 km no início (dilui orçamento)",
        ),
        (
            "Faixa etária",
            "25–54 (core 30–49)",
            "Não forçar idade no Search; no Display 25–54",
            "25–54 cobre comprador/vendedor de carro",
        ),
        (
            "Gênero",
            "Todos",
            "Todos",
            "Sem restrição",
        ),
        (
            "Interesses / palavras",
            "Carros, DETRAN, compra/venda veículo, moto",
            "despachante cosmópolis, transferência veículo, licenciamento",
            "Google = intenção; Meta = contexto + remarketing",
        ),
        (
            "Horário",
            "Seg–Sáb 8h–20h (pico 10h–13h e 17h–20h)",
            "Mesmo horário + finais de semana se atender",
            "Alinhar com horário real de WhatsApp",
        ),
        (
            "Criativo",
            "Fotos pessoas/carro + CTA WhatsApp + telefone",
            "Extensões de ligação/local + sitelinks serviços",
            "Usar os criativos já feitos no repositório",
        ),
        (
            "KPI principal",
            "Conversas WhatsApp → fechamentos",
            "Cliques → ligações/WhatsApp → fechamentos",
            "Marcar origem no WhatsApp (Meta/Google)",
        ),
    ]
    for i, row in enumerate(data):
        for c, val in enumerate(row, 1):
            cell = ws.cell(4 + i, c, val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[4 + i].height = 48

    r = 4 + len(data) + 2
    ws.cell(r, 1, "SPLIT DE ORÇAMENTO SUGERIDO").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    for i, h in enumerate(["Investimento", "% Meta", "% Google", "R$ Meta", "R$ Google", "Por quê"], 1):
        style_header(ws.cell(r, i, h), "C4840A")
    for inv, pm, pg, why in [
        (500, 0.70, 0.30, "Pouca verba: gerar conversa no WhatsApp primeiro"),
        (1000, 0.60, 0.40, "Escala: capturar busca + reforçar marca"),
    ]:
        r += 1
        ws.cell(r, 1, inv)
        style_money(ws.cell(r, 1))
        style_pct(ws.cell(r, 2, pm))
        style_pct(ws.cell(r, 3, pg))
        style_money(ws.cell(r, 4, inv * pm))
        style_money(ws.cell(r, 5, inv * pg))
        ws.cell(r, 6, why).font = Font(name="Arial", size=10)

    r += 3
    ws.cell(r, 1, "RAIO DETALHADO").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    for line in [
        "Pino: R. Dr. Luiz Nicolau Nolandi / Campo do Kigol — Novo Horizonte, Cosmópolis-SP (CEP 13150-548)",
        "Raio fase 1 (mês 1–2): 15 km — Cosmópolis inteira + borda Artur Nogueira",
        "Raio fase 2 (se CPA ok): 20–25 km — Artur Nogueira e entorno",
        "Evitar: Campinas/Paulínia amplo no começo (CPA sobe e atendimento dilui)",
    ]:
        ws.cell(r, 1, line).font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    autosize(ws, [22, 38, 38, 40, 14, 42])


def sheet_dre(wb, orcamento, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"DRE projetado — investimento R$ {orcamento:,.0f}/mês em tráfego".replace(",", ".")
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3358")
    ws.merge_cells("A1:H1")
    ws["A2"] = "Lucro após mídia = (Vendas × Lucro bruto médio) − Investimento em ads. Não inclui aluguel/salário fixo."
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="5A6D86")

    # Três cenários lado a lado
    ws["A4"] = "CENÁRIOS DE CPA (custo por serviço fechado)"
    ws["A4"].font = Font(name="Arial", bold=True, size=11, color="1A3358")

    headers = ["Indicador", "Conservador", "Base", "Otimista"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(5, i, h))

    cens = {
        "Conservador": montar_cenario(orcamento, "conservador"),
        "Base": montar_cenario(orcamento, "base"),
        "Otimista": montar_cenario(orcamento, "otimista"),
    }
    lines = [
        ("Investimento mídia", "orcamento", "money"),
        ("Verba Meta", "spend_meta", "money"),
        ("Verba Google", "spend_google", "money"),
        ("CPA Meta", "cpa_meta", "money"),
        ("CPA Google", "cpa_google", "money"),
        ("Vendas Meta", "vendas_meta", "num"),
        ("Vendas Google", "vendas_google", "num"),
        ("Vendas totais / mês", "vendas", "num"),
        ("Vendas p/ break-even", "be_vendas", "num"),
        ("Receita bruta (ticket médio)", "receita", "money"),
        ("(-) Custos processos/op. (implícitos no lucro)", None, "note"),
        ("Lucro bruto dos serviços", "lucro_bruto", "money"),
        ("(-) Investimento mídia", "orcamento", "money"),
        ("= Lucro após mídia", "lucro_apos_midia", "money"),
        ("ROI (lucro após mídia / invest.)", "roi", "pct"),
        ("ROAS lucro (lucro bruto / invest.)", "roas_lucro", "pct"),
    ]

    for i, (label, key, kind) in enumerate(lines):
        rr = 6 + i
        ws.cell(rr, 1, label).font = Font(name="Arial", bold=("=" in label or "Lucro após" in label), size=10)
        for j, nome in enumerate(["Conservador", "Base", "Otimista"], 2):
            if key is None:
                ws.cell(rr, j, "—").font = Font(name="Arial", size=10, color="5A6D86")
                continue
            val = cens[nome][key]
            cell = ws.cell(rr, j, val)
            if kind == "money":
                style_money(cell)
            elif kind == "pct":
                style_pct(cell)
            else:
                cell.number_format = "0.00"
                cell.font = Font(name="Arial", size=10)
            if "Lucro após" in label:
                color = "1A7F37" if val >= 0 else "B42318"
                cell.font = Font(name="Arial", bold=True, color=color, size=10)

    # DRE mensal 6 meses — cenário base com curva de aprendizado
    r0 = 6 + len(lines) + 2
    ws.cell(r0, 1, "DRE MENSAL 6 MESES — CENÁRIO BASE (com curva de aprendizado)").font = Font(
        name="Arial", bold=True, size=12, color="1A3358"
    )
    ws.cell(r0 + 1, 1, "Mês 1: CPA 20% pior | Mês 2: CPA base | Mês 3+: CPA 10% melhor (otimização)").font = Font(
        name="Arial", italic=True, size=9, color="5A6D86"
    )

    headers2 = ["Mês", "Investimento", "Vendas", "Receita", "Lucro bruto serviços", "Lucro após mídia", "ROI", "Lucro acumulado", "Status"]
    for i, h in enumerate(headers2, 1):
        style_header(ws.cell(r0 + 3, i, h), "C4840A")

    # CPA factors by month
    factors = [1.20, 1.00, 0.90, 0.90, 0.85, 0.85]
    acum = 0.0
    first_roi_month = None
    be_month = None
    chart_start = r0 + 4

    for m, fac in enumerate(factors, 1):
        # rebuild with adjusted CPA by scaling vendas inversely
        base = montar_cenario(orcamento, "base")
        vendas = base["vendas"] / fac
        receita = vendas * TICKET_MEDIO
        lucro_b = vendas * LUCRO_MEDIO
        lucro_m = lucro_b - orcamento
        roi = lucro_m / orcamento
        acum += lucro_m
        if be_month is None and acum >= 0:
            be_month = m
        if first_roi_month is None and lucro_m > 0:
            first_roi_month = m
        status = "Break-even acumulado" if acum >= 0 and (be_month == m) else ("ROI positivo no mês" if lucro_m > 0 else "Abaixo do BE")
        rr = r0 + 3 + m
        ws.cell(rr, 1, f"Mês {m}").font = Font(name="Arial", size=10)
        style_money(ws.cell(rr, 2, orcamento))
        ws.cell(rr, 3, vendas).number_format = "0.0"
        style_money(ws.cell(rr, 4, receita))
        style_money(ws.cell(rr, 5, lucro_b))
        cell_l = ws.cell(rr, 6, lucro_m)
        style_money(cell_l)
        cell_l.font = Font(name="Arial", bold=True, color="1A7F37" if lucro_m >= 0 else "B42318", size=10)
        style_pct(ws.cell(rr, 7, roi))
        style_money(ws.cell(rr, 8, acum))
        ws.cell(rr, 9, status).font = Font(name="Arial", size=9)

    # LT summary
    r = chart_start + 7
    ws.cell(r, 1, "LT (PRAZO) — LEITURA").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    ws.cell(r, 1, f"LT break-even (vendas no mês): {base['be_vendas']:.1f} serviços fechados para cobrir R$ {orcamento:.0f}").font = Font(name="Arial", size=10)
    r += 1
    ws.cell(r, 1, f"LT break-even acumulado (cenário base com curva): Mês {be_month or '—'}").font = Font(name="Arial", size=10)
    r += 1
    ws.cell(r, 1, f"LT que começa a dar ROI no mês (lucro após mídia > 0): Mês {first_roi_month or '—'}").font = Font(name="Arial", size=10)
    r += 1
    ws.cell(r, 1, "ROI consistente (após otimização CPA): a partir do Mês 2–3, mantendo tracking de origem no WhatsApp.").font = Font(name="Arial", size=10)

    # chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Lucro após mídia por mês"
    chart.y_axis.title = "R$"
    data = Reference(ws, min_col=6, min_row=r0 + 3, max_row=r0 + 9)
    cats = Reference(ws, min_col=1, min_row=r0 + 4, max_row=r0 + 9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "A" + str(r + 3))

    line = LineChart()
    line.title = "Lucro acumulado"
    line.y_axis.title = "R$"
    data2 = Reference(ws, min_col=8, min_row=r0 + 3, max_row=r0 + 9)
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats)
    line.width = 15
    line.height = 8
    line.style = 10
    ws.add_chart(line, "E" + str(r + 3))

    autosize(ws, [48, 14, 12, 14, 20, 16, 12, 16, 22])
    return be_month, first_roi_month


def sheet_breakeven(wb):
    ws = wb.create_sheet("05_BreakEven_ROI")
    ws["A1"] = "Break-even e ROI — quanto precisa vender"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3358")

    headers = ["Investimento", "Lucro médio/serviço", "Vendas p/ BE", "Se fechar 4/mês", "Se fechar 6/mês", "Se fechar 8/mês", "LT BE (se 4 vendas/mês)"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h))

    for row_i, inv in enumerate([500, 1000], 4):
        be = inv / LUCRO_MEDIO
        ws.cell(row_i, 1, inv)
        style_money(ws.cell(row_i, 1))
        style_money(ws.cell(row_i, 2, LUCRO_MEDIO))
        ws.cell(row_i, 3, be).number_format = "0.00"
        for col, vendas in enumerate([4, 6, 8], 4):
            lucro = vendas * LUCRO_MEDIO - inv
            cell = ws.cell(row_i, col, lucro)
            style_money(cell)
            cell.font = Font(name="Arial", bold=True, color="1A7F37" if lucro >= 0 else "B42318", size=10)
        # months to BE if only 4 sales/month contribution
        contrib = 4 * LUCRO_MEDIO - inv
        if contrib >= 0:
            lt = 1
        else:
            # shouldn't happen with these numbers
            lt = max(1, inv / (4 * LUCRO_MEDIO))
        ws.cell(row_i, 7, f"{lt:.0f} mês" if lt <= 1 else f"{lt:.1f} meses").font = Font(name="Arial", size=10)

    r = 8
    ws.cell(r, 1, "POR SERVIÇO — quantos precisa para pagar a mídia").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    r += 1
    for i, h in enumerate(["Serviço", "Lucro unitário", "Qtd p/ BE R$500", "Qtd p/ BE R$1.000"], 1):
        style_header(ws.cell(r, i, h), "C4840A")
    for nome, *_, lb in SERVICOS:
        r += 1
        ws.cell(r, 1, nome).font = Font(name="Arial", size=10)
        style_money(ws.cell(r, 2, lb))
        ws.cell(r, 3, 500 / lb).number_format = "0.00"
        ws.cell(r, 4, 1000 / lb).number_format = "0.00"

    r += 3
    ws.cell(r, 1, "INTERPRETAÇÃO LT").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    notes = [
        "LT break-even (operacional): quantidade de serviços no mês cujo lucro bruto cobre 100% da mídia.",
        f"Com mix atual (~R$ {LUCRO_MEDIO:.0f}/serviço): ~{500/LUCRO_MEDIO:.1f} vendas em R$500 e ~{1000/LUCRO_MEDIO:.1f} em R$1.000.",
        "LT break-even acumulado: mês em que o lucro após mídia acumulado fica ≥ 0 (na aba DRE, com curva de aprendizado).",
        "LT início de ROI: primeiro mês com lucro após mídia > 0 de forma recorrente — tipicamente Mês 1 (base) ou Mês 2 (se CPA alto no começo).",
        "ROI consistente: a partir do Mês 2–3, quando criativos, públicos e palavras-chave já foram otimizados.",
        "Acompanhe no WhatsApp tags: #meta / #google e serviço fechado — sem isso o CPA fica 'achismo'.",
    ]
    for n in notes:
        r += 1
        ws.cell(r, 1, "• " + n).font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    autosize(ws, [34, 16, 16, 16, 16, 16, 22])


def sheet_plano_acao(wb):
    ws = wb.create_sheet("06_Plano_30_dias")
    ws["A1"] = "Plano prático — primeiros 30 dias"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3358")
    headers = ["Semana", "Ação", "Canal", "Budget sugerido (se R$500)", "Budget sugerido (se R$1.000)", "Meta"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(3, i, h))
    plan = [
        ("Semana 1", "Subir Pixel/API WhatsApp + campanha mensagens + 3 criativos", "Meta", 350, 600, "30–50 conversas"),
        ("Semana 1", "Campanha Search: despachante Cosmópolis + transferência", "Google", 150, 400, "15–30 cliques qualificados"),
        ("Semana 2", "Cortar anúncio com CPL alto; reforçar melhor criativo", "Meta", 350, 600, "CPA conversa ↓ 20%"),
        ("Semana 2", "Ativar extensão de local + ligação", "Google", 150, 400, "Chamadas/rotas Maps"),
        ("Semana 3", "Remarketing quem abriu WhatsApp e não fechou", "Meta", 350, 550, "+fechamentos"),
        ("Semana 3", "Ampliar keywords licenciamento / placa", "Google", 150, 450, "Mais intenção"),
        ("Semana 4", "Fechar DRE real: gasto × serviços × lucro", "Os dois", 500, 1000, "Decidir manter/escalar"),
    ]
    for i, row in enumerate(plan):
        for c, v in enumerate(row, 1):
            cell = ws.cell(4 + i, c, v)
            cell.font = Font(name="Arial", size=10)
            if c in (4, 5) and isinstance(v, (int, float)):
                style_money(cell)
    r = 12
    ws.cell(r, 1, "Checklist tracking").font = Font(name="Arial", bold=True, size=12, color="1A3358")
    for item in [
        "☐ Tag no WhatsApp: origem Meta ou Google",
        "☐ Planilha diária: data | canal | serviço | valor | lucro",
        "☐ Meta: campanha só WhatsApp (não só curtida)",
        "☐ Google: localização presença / raio 15–20 km",
        "☐ Usar criativos da pasta /criativos do site",
        "☐ Site na bio + link Google Meu Negócio",
    ]:
        r += 1
        ws.cell(r, 1, item).font = Font(name="Arial", size=10)

    autosize(ws, [12, 55, 12, 22, 24, 24])


def main():
    wb = Workbook()
    sheet_capa(wb)
    sheet_premissas(wb)
    sheet_canais(wb)
    sheet_dre(wb, 500, "03_DRE_R$500")
    sheet_dre(wb, 1000, "04_DRE_R$1000")
    sheet_breakeven(wb)
    sheet_plano_acao(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("OK", OUT)
    print(f"Lucro medio={LUCRO_MEDIO:.2f} Ticket={TICKET_MEDIO:.2f}")
    for inv in (500, 1000):
        b = montar_cenario(inv, "base")
        print(inv, "vendas", round(b["vendas"], 1), "lucro_midia", round(b["lucro_apos_midia"], 1), "BE", round(b["be_vendas"], 1))


if __name__ == "__main__":
    main()
